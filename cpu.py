from bs4 import BeautifulSoup
import bs4
import requests
import openpyxl

url: str = "https://www.overclockers.ua/cpu/info/"
page: requests.Response = requests.get(url)

soup: BeautifulSoup = BeautifulSoup(page.text, "lxml")

wb: openpyxl.Workbook = openpyxl.reader.excel.load_workbook(filename="cpu_start.xlsx", data_only=True)
wb.active = 0
ws: list = wb.active

ultagamd: bs4.element.Tag = soup.find(id="spec_wrapper").find('ul', id="amd")

n: int = 2

item: bs4.element.Tag
for item in ultagamd.find_all('li'):
    if str(item) != '<li class="first switch"><a href="">показать все</a> ▼</li>':
        ws['A' + str(n)] = item.text
        ws['B' + str(n)] = item.find('a').attrs['href']
        lurl: str = "https://www.overclockers.ua" + str(item.find('a').attrs['href'])
        lpage: requests.Response = requests.get(lurl)
        lsoup: BeautifulSoup = BeautifulSoup(lpage.text, "lxml")
        tbl: bs4.element.Tag = lsoup.find('div', id="spec_table")
        litem: bs4.element.Tag
        for litem in tbl.find_all('tr'):
            if litem.find('td', class_="gr3") != None:
                llitem: bs4.element.Tag = litem.find_all('td')[1]
                if litem.find('td', class_="gr3").text == "Ядро":
                    ws['C' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Количество ядер":
                    ws['D' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Количество потоков":
                    ws['E' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Техпроцесс, нм":
                    ws['F' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Разъем":
                    ws['G' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Частота, МГц":
                    ws['H' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Множитель":
                    ws['I' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Тип памяти":
                    ws['J' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "кэш L1, КБ":
                    ws['K' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "кэш L2, КБ":
                    ws['L' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "кэш L3, КБ":
                    ws['M' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "TDP, Вт":
                    ws['N' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Предельная температура, °C":
                    ws['O' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Дата выпуска":
                    ws['P' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Стоимость, $":
                    ws['Q' + str(n)] = llitem.text
        n += 1

ultagintel: bs4.element.Tag = soup.find(id="spec_wrapper").find('ul', id="intel")

item: bs4.element.Tag
for item in ultagintel.find_all('li'):
    if str(item) != '<li class="first switch"><a href="">показать все</a> ▼</li>':
        ws['A' + str(n)] = item.text
        ws['B' + str(n)] = item.find('a').attrs['href']
        lurl: str = "https://www.overclockers.ua" + str(item.find('a').attrs['href'])
        lpage: requests.Response = requests.get(lurl)
        lsoup: BeautifulSoup = BeautifulSoup(lpage.text, "lxml")
        tbl: bs4.element.Tag = lsoup.find('div', id="spec_table")
        for litem in tbl.find_all('tr'):
            if litem.find('td', class_="gr3") != None:
                llitem: bs4.element.Tag = litem.find_all('td')[1]
                if litem.find('td', class_="gr3").text == "Ядро":
                    ws['C' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Количество ядер":
                    ws['D' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Количество потоков":
                    ws['E' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Техпроцесс, нм":
                    ws['F' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Разъем":
                    ws['G' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Частота, МГц":
                    ws['H' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Множитель":
                    ws['I' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Тип памяти":
                    ws['J' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "кэш L1, КБ":
                    ws['K' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "кэш L2, КБ":
                    ws['L' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "кэш L3, КБ":
                    ws['M' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "TDP, Вт":
                    ws['N' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Предельная температура, °C":
                    ws['O' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Дата выпуска":
                    ws['P' + str(n)] = llitem.text
                if litem.find('td', class_="gr3").text == "Стоимость, $":
                    ws['Q' + str(n)] = llitem.text
        n += 1

wb.save("cpu.xlsx")