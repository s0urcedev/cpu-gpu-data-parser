from bs4 import BeautifulSoup
import bs4
import requests
import openpyxl

url: str = "https://www.overclockers.ua/video/gpu/"
page: requests.Response = requests.get(url)

soup: BeautifulSoup = BeautifulSoup(page.text, "lxml")

wb: openpyxl.Workbook = openpyxl.reader.excel.load_workbook(filename="gpu_start.xlsx", data_only=True)
wb.active = 0
ws: list = wb.active

ultagamd: bs4.element.Tag = soup.find(id="spec_wrapper").find('ul', id="amd")

n: int = 2

item: bs4.element.Tag
for item in ultagamd.find_all('li'):
	if str(item) != '<li class="first switch"><a href="">показать все</a> ▼</li>':
		ws['A' + str(n)] = item.text
		ws['B' + str(n)] = item.find('a').attrs['href']
		lurl: str = "https://www.overclockers.ua" + str(item.find('a').attrs['href'])
		lpage: requests.Response = requests.get(lurl)
		lsoup: BeautifulSoup = BeautifulSoup(lpage.text, "lxml")
		tbl: bs4.element.Tag = lsoup.find('div', id="spec_table")
		litem: bs4.element.Tag
		for litem in tbl.find_all('tr'):
			if litem.find('td', class_="gr3") != None:
				llitem: bs4.element.Tag = litem.find_all('td')[1]
				if litem.find('td', class_="gr3").text == "Ядро":
					ws['C' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Техпроцесс, нм":
					ws['D' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Частота работы ядра, МГц":
					ws['E' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Частота работы шейдерных блоков, МГц":
					ws['F' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Шейдерных блоков":
					ws['G' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Частота работы памяти (DDR), МГц":
					ws['H' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Шина памяти":
					ws['I' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Объем памяти":
					ws['J' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "DirectX":
					ws['K' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Интерфейс":
					ws['L' + str(n)] = llitem.text
		n += 1

ultagnvidia: bs4.element.Tag = soup.find(id="spec_wrapper").find('ul', id="nvidia")

item: bs4.element.Tag
for item in ultagnvidia.find_all('li'):
	if str(item) != '<li class="first switch"><a href="">показать все</a> ▼</li>':
		ws['A' + str(n)] = item.text
		ws['B' + str(n)] = item.find('a').attrs['href']
		lurl: str = "https://www.overclockers.ua" + str(item.find('a').attrs['href'])
		lpage: requests.Response = requests.get(lurl)
		lsoup: BeautifulSoup = BeautifulSoup(lpage.text, "lxml")
		tbl: bs4.element.Tag = lsoup.find('div', id="spec_table")
		for litem in tbl.find_all('tr'):
			if litem.find('td', class_="gr3") != None:
				llitem: bs4.element.Tag = litem.find_all('td')[1]
				if litem.find('td', class_="gr3").text == "Ядро":
					ws['C' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Техпроцесс, нм":
					ws['D' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Частота работы ядра, МГц":
					ws['E' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Частота работы шейдерных блоков, МГц":
					ws['F' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Шейдерных блоков":
					ws['G' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Частота работы памяти (DDR), МГц":
					ws['H' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Шина памяти":
					ws['I' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Объем памяти":
					ws['J' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "DirectX":
					ws['K' + str(n)] = llitem.text
				if litem.find('td', class_="gr3").text == "Интерфейс":
					ws['L' + str(n)] = llitem.text
		n += 1

wb.save("gpu.xlsx")